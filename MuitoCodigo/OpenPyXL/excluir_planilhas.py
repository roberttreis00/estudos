import openpyxl

wb = openpyxl.load_workbook('testes1.xlsx')
plan = wb['mercado']
colunas = wb.sheetnames
print(colunas)

wb.copy_worksheet(plan)
wb.save('testes2.xlsx')

# sheet para excluir
sheet_excluir = wb['mercado Copy']
wb.remove(sheet_excluir)  # primeira opcao e posterior a segunda opcao ambas funciona
del wb['mercado Copy2']
del wb['mercado Copy3']
del wb['mercado Copy1']

wb.save('testes2.xlsx')

colunas = wb.sheetnames
print(colunas)
