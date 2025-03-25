from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active

# aplicar cor
red_font = Font(color='FF0000')
green_font = Font(color='00FF00')

amarelo_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

ws['A2'].font = red_font
ws['A2'].fill = amarelo_fill
ws['A2'] = 'Texto vermelho e fundo amarelo'

wb.save('estiloCOR.xlsx')
