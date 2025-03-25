from openpyxl import Workbook
from openpyxl.styles import Font

# Criar nova planilha
wb = Workbook()
ws = wb.active

# aplicar formatações
bold_font = Font(bold=True)
italic_font = Font(italic=True)
bold_italic_font = Font(bold=True, italic=True)

# aplicar formatação em especificas celulas
ws['A1'].font = bold_font
ws['A1'] = 'Fonte negrito'

ws['B1'].font = italic_font
ws['B1'] = 'Fonte em italico'

ws['C1'].font = bold_italic_font
ws['C1'] = 'Negrito e italico texto'

wb.save('style_exemplos.xlsx')
