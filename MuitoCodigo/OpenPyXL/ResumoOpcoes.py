import openpyxl

# Cria uma nova planilha
wb = openpyxl.Workbook()
# wb.create_sheet()  # cria outra planilha na mesma planilha da para nomear e colocar os indexs

wb.save('planilha1.xlsx')
# -- -- -- -- --
# Ler planilha/ Primeiro selecio qual planilha é depois pode acessar os valores de cada celula
wb2 = openpyxl.load_workbook('testes2.xlsx')
ws = wb2['mercado teste']

print(ws['A1'].value)
