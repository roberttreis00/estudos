from pprint import pprint
import openpyxl

ws = openpyxl.load_workbook('MegaPlanilha.xlsx')['MEGA SENA']


def acessando_todos_valores_coluna(planilha, index_coluna):
    qtd_linhas = len([dado for dado in planilha.values])
    coluna = True
    data = dict()

    for index in range(1, qtd_linhas+1):
        if coluna:
            nome_coluna = planilha.cell(row=index, column=index_coluna).value
            data[nome_coluna] = []

            coluna = False
            continue

        dado = planilha.cell(row=index, column=index_coluna).value
        data[nome_coluna].append(dado)

    return data


print(acessando_todos_valores_coluna(ws, 1))
print(acessando_todos_valores_coluna(ws, 3))
